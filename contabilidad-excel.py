import openpyxl

def modificar_filas(archivo_entrada, archivo_salida, nombre_hoja):
    # Cargar el archivo Excel
    wb = openpyxl.load_workbook(archivo_entrada)
    sheet = wb[nombre_hoja]

    # Obtener el número de filas en la hoja
    num_filas = sheet.max_row

    # Variables para rastrear los últimos valores encontrados
    ultimo_asto = None
    ultimo_concepto = None
    ultimo_fecha = None

    # Recorrer filas del archivo de entrada
    for i in range(2, num_filas + 1):  # Empezar desde la segunda fila
        asto = sheet.cell(row=i, column=2).value  # Obtener el valor de "Asto"
        concepto = sheet.cell(row=i, column=7).value  # Obtener el valor de "Concepto"
        fecha = sheet.cell(row=i, column=1).value  # Obtener el valor de "Fecha"

        # Modificar el valor de "Asto" si es igual al último valor de "Asto" encontrado
        if asto == ultimo_asto:
            sheet.cell(row=i, column=2, value="")

        # Modificar el valor de "Concepto" si es igual al último valor de "Concepto" encontrado
        if concepto == ultimo_concepto:
            sheet.cell(row=i, column=7, value="")

        # Modificar el valor de "Fecha" si es igual al último valor de "Fecha" encontrado
        if fecha == ultimo_fecha:
            sheet.cell(row=i, column=1, value="")
       
        # Actualizar los últimos valores encontrados
        ultimo_asto = asto
        ultimo_concepto = concepto
        ultimo_fecha = fecha

    # Guardar el archivo modificado
    wb.save(archivo_salida)

# Ejemplo de uso
archivo_entrada = 'Muestra diario de movimientos oficial 2023.xlsx'
archivo_salida = 'archivo_resultado.xlsx'
nombre_hoja = 'Diario de movimientos oficial'  # Nombre de la hoja donde quieres realizar la operación
modificar_filas(archivo_entrada, archivo_salida, nombre_hoja)
