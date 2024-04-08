# import openpyxl
# #este codigo es para añadir filas extras...

# def agregar_filas_extra(archivo_entrada, archivo_salida, nombre_hoja):
#     # Cargar el archivo Excel
#     wb = openpyxl.load_workbook(archivo_entrada)
    
#     # Seleccionar la hoja específica por su nombre
#     sheet = wb[nombre_hoja]

#     # Crear una nueva hoja para el archivo de salida
#     wb_salida = openpyxl.Workbook()
#     sheet_salida = wb_salida.active
    
#     # Definir la columna de "nombre de tienda"
#     columna_nombre_tienda = None

#     # Buscar la columna "nombre de tienda" y agregar filas adicionales debajo de cada celda
#     for col_idx, col in enumerate(sheet.iter_cols(), start=1):
#         for cell in col:
#             if cell.value == "Nombre de la tienda":
#                 columna_nombre_tienda = col_idx
#                 break
#         if columna_nombre_tienda:
#             break

#     if columna_nombre_tienda:
#         for row_idx, row in enumerate(sheet.iter_rows(), start=1):
#             nombre_tienda = row[columna_nombre_tienda - 1].value
#             sheet_salida.append([cell.value for cell in row])
#             if nombre_tienda == "Nombre de la tienda":
#                 continue
#             for i in range(7):
#                 sheet_salida.append([''] * sheet.max_column)

#     # Guardar el archivo de salida
#     wb_salida.save(archivo_salida)


# archivo_entrada = 'Archivo-pedidos-Natulim-script.xlsx'
# archivo_salida = 'archivo_salida.xlsx'
# nombre_hoja = 'Sheet1'
# agregar_filas_extra(archivo_entrada, archivo_salida, nombre_hoja)
import openpyxl

def agregar_filas_extra(archivo_entrada, archivo_salida, nombre_hoja):
    # Cargar el archivo Excel
    wb = openpyxl.load_workbook(archivo_entrada)
    
    # Seleccionar la hoja específica por su nombre
    sheet = wb[nombre_hoja]

    # Crear una nueva hoja para el archivo de salida
    wb_salida = openpyxl.Workbook()
    sheet_salida = wb_salida.active
    
    # Definir las columnas de interés
    columna_nombre_tienda = None
    columna_sku = 5  # Columna E
    columna_numero_cajas = 6  # Columna F
    columna_precio = 7  # Columna G
    columna_precio_original = 8  # Columna H
    columna_sku_original = 9  # Columna I
    columna_numero_cajas_original = 10  # Columna J

    # Buscar la columna "nombre de tienda" y agregar filas adicionales debajo de cada celda
    for col_idx, col in enumerate(sheet.iter_cols(), start=1):
        for cell in col:
            if cell.value == "Nombre de la tienda":
                columna_nombre_tienda = col_idx
                break
        if columna_nombre_tienda:
            break

    if columna_nombre_tienda:
        for row_idx, row in enumerate(sheet.iter_rows(), start=1):
            nombre_tienda = row[columna_nombre_tienda - 1].value
            sheet_salida.append([cell.value for cell in row])
            if nombre_tienda == "Nombre de la tienda":
                continue
            for i in range(7):
                sheet_salida.append([''] * sheet.max_column)

                # Check if columns G, I, J have values and move them to E, F, G
                if all(row[columna].value is not None for columna in (columna_precio, columna_precio_original, columna_sku_original)):
                    precio = row[columna_precio].value
                    sku = row[columna_sku_original].value
                    numero_cajas = row[columna_numero_cajas_original].value
                    sheet_salida.cell(row=row_idx + 7, column=columna_sku).value = sku
                    sheet_salida.cell(row=row_idx + 7, column=columna_numero_cajas).value = numero_cajas
                    sheet_salida.cell(row=row_idx + 7, column=columna_precio).value = precio

    # Guardar el archivo de salida
    wb_salida.save(archivo_salida)
# archivo entrada es el archivo principal que queremos manipular
archivo_entrada = 'Test (4).xlsx'
# el archivo que nos da el Script
archivo_salida = 'archivo_salida-columnas.xlsx'
nombre_hoja = 'Sheet1'
agregar_filas_extra(archivo_entrada, archivo_salida, nombre_hoja)
